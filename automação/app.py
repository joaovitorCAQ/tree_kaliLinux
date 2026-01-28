import os 
import sys 
import shutil
from openpyxl import load_workbook , Workbook # biblioteca para acessar
planilha = load_workbook('vendas_de_lanches.xlsx') # nome do arquivo exel precisar dentro para abrir com loar_workbook
pagina_vendas = planilha['Sheet'] # nome da primeira pagina que esta la dentro do exel 

# iter_rows permite entrar em todas as linhas da sua planilha (posso colocar a linha inicial ou final que eu eu quero comecar ou terminar oque quiser)
for linha,i in enumerate(pagina_vendas.iter_rows(values_only=True)): # values_only retorna apenas os valores da celula
    print(linha,i)

# segunda parte(criar nova planilha do zero)
planilha_contas = Workbook()
pagina1 = planilha_contas.active

with open('Fornecedor,Descrição,Valor,Data de.txt','r',encoding='utf-8') as arquivo :
    for linha in arquivo:
        pagina1.append(linha.split(","))
planilha_contas.save('contas a pagar.xlsx')

def main():
    print("Planilha lida com sucesso!")

if __name__ == "__main__":
    main()